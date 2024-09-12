import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, numbers

def process_csv_to_xlsx_all_pin(digital_file_path, time_increment=5):
    digital_df = pd.read_csv(digital_file_path)

    # 새로운 엑셀 파일 생성
    workbook = Workbook()
    sheet = workbook.active

    # 첫 번째 행을 빈 행으로 두기 위해서 빈 리스트 추가
    sheet.append([''] * (len(digital_df.columns) + 2))

    # 두 번째 행에 헤더 추가 (C~Q 열에 해당하는 헤더 추가)
    gpio_custom_headers = ['Time[s]', 'GPIO0', 'GPIO1', 'GPIO2', 'GPIO3', 'GPIO4', 'GPIO5', 'GPIO8', 'GPIO9', 'GPIO10', 'GPIO11', 'GPIO12', 'GPIO13', 'GPIO19', 'GPIO26', 'GPIO27', 'GPIO28']  # C~Q 열에 해당

    headers = [''] * 2 + gpio_custom_headers
    sheet.append(headers)

    sheet['T2'] = 'Time[ns]'
    sheet['U2'] = 'TotalTime[ns]'
    sheet['V2'] = 'Line num'

    gpio_test_row = None  # 첫 번째 조건을 만족하는 행
    peri_test_row = len(digital_df) + 3 - 1  # 두 번째 조건을 만족하는 행
    peri_test_count = 0   # peri_test_row를 찾기 위한 카운트, 초기화 필요
    condition1_count = 0  # 조건 1이 해당된 횟수를 카운트하는 변수

    for i, row in enumerate(digital_df.itertuples(index=False), start=3):
        # 각 셀 값을 시트에 입력
        for j in range(2, 19):
            sheet.cell(row=i, column=j+1, value=row[j-2])

        # D열이 1이고 E열에서 P열이 0인 첫 번째 행 찾기 (gpio_test_row)
        if gpio_test_row is None:
            if sheet.cell(row=i, column=4).value == 1 and all(sheet.cell(row=i, column=j).value == 0 for j in range(5, 17)):
                gpio_test_row = i

        # gpio_test_row 이후에 D열이 0이고 E열이 1이며 F열에서 P열이 0인 행을 카운트해서 두 번째 행을 찾기 (peri_test_row)
        if gpio_test_row is not None and i > gpio_test_row:
            if (sheet.cell(row=i, column=4).value == 0 and 
                sheet.cell(row=i, column=5).value == 1 and 
                all(sheet.cell(row=i, column=j).value == 0 for j in range(6, 17))):
                peri_test_count += 1  # 조건을 만족할 때마다 카운트 증가
                if peri_test_count == 2:  # 두 번째로 해당하는 행을 peri_test_row에 저장
                    peri_test_row = i

    last_row = len(digital_df) + 3 - 1  # 데이터가 끝나는 행 (CSV의 행 수 + 3번째 행부터 시작)

    # 기존 로직: 4행부터 마지막 행까지 반복하면서 데이터를 처리
    for i in range(4, last_row):
        # T열 수식: (다음 행의 C열 값 - 현재 행의 C열 값) * 1000000000
        next_row = i + 1
        sheet[f'T{i}'] = f'=(C{next_row} - C{i}) * 1000000000'
        sheet[f'T{i}'].number_format = numbers.FORMAT_NUMBER  # T열 숫자 형식 설정

        # U열 수식
        if i == 4:
            sheet[f'U{i}'] = 0  # 4행의 U열에는 숫자 0을 입력
        else:
            sheet[f'U{i}'] = f'=(U{i-1} + T{i-1})'
        sheet[f'U{i}'].number_format = numbers.FORMAT_NUMBER  # U열 숫자 형식 설정
        
        # V열 수식: U열 값을 time_increment로 나눔
        if i == 4:
            sheet[f'V{i}'] = 0  # 4행의 V열에는 숫자 0을 입력
        else:
            sheet[f'V{i}'] = f'=U{i} / {time_increment}'
        sheet[f'V{i}'].number_format = numbers.FORMAT_NUMBER  # V열 숫자 형식 설정

        # D열에서 I열까지의 값
        di_values = [sheet.cell(row=i, column=col).value for col in range(4, 10)]  # D(4) ~ I(9)
        # J열에서 O열까지의 값
        jo_values = [sheet.cell(row=i, column=col).value for col in range(10, 16)]  # J(10) ~ O(15)
        # Q열에서 S열까지의 값
        qs_values = [sheet.cell(row=i, column=col).value for col in range(17, 20)]  # Q(17) ~ S(19)

        # 조건 1: D에서 I열까지 값 중 1이 있고 J에서 O열이 모두 0이고 Q에서 S열이 모두 1일 경우 빨간색
        if any(value == 1 for value in di_values) and jo_values.count(1) == 0 and any(value == 1 for value in qs_values):
            condition1_count += 1  # 조건 1에 해당되면 카운트 증가
            sheet[f'U{i}'].font = Font(color="FF0000")
            sheet[f'V{i}'].font = Font(color="FF0000")

        # 조건 2: D에서 I열까지 값이 모두 0이고 J에서 O열 값 중 1이 있고 Q에서 S열이 모두 1일 경우 파란색
        elif di_values.count(1) == 0 and any(value == 1 for value in jo_values) and any(value == 1 for value in qs_values):
            sheet[f'U{i}'].font = Font(color="0000FF")
            sheet[f'V{i}'].font = Font(color="0000FF")

        # 조건 3: 이전 행이 조건 1에 해당하고, peri_test_row 이후에 D열에서 I열, J열에서 O열 모두 0이고 Q에서 S열이 1일 경우 빨간색
        #if peri_test_row is not None and i > peri_test_row:
        prev_row = i - 1
        prev_di_values = [sheet.cell(row=prev_row, column=col).value for col in range(4, 10)]
        prev_jo_values = [sheet.cell(row=prev_row, column=col).value for col in range(10, 16)]
        prev_qs_values = [sheet.cell(row=prev_row, column=col).value for col in range(17, 20)]

        if any(value == 1 for value in prev_di_values) and prev_jo_values.count(1) == 0 and any(value == 1 for value in prev_qs_values):
            if di_values.count(1) == 0 and jo_values.count(1) == 0 and any(value == 1 for value in prev_qs_values):
                sheet[f'U{i}'].font = Font(color="FF0000")
                sheet[f'V{i}'].font = Font(color="FF0000")

        # 조건 4: i가 gpio_test_row와 peri_test_row 사이의 값이고, D에서 S열 중 1이 하나만 있으면 파란색
        if gpio_test_row is not None and peri_test_row is not None and gpio_test_row < i < peri_test_row:
            ds_values = [sheet.cell(row=i, column=col).value for col in range(4, 20)]  # D(4) ~ S(19)
            if ds_values.count(1) == 1:  # D에서 S열 중 1이 하나만 있으면
                sheet[f'U{i}'].font = Font(color="0000FF")
                sheet[f'V{i}'].font = Font(color="0000FF")

    destination_file = digital_file_path.replace('.csv', '.xlsx')
    try:
        workbook.save(destination_file)
        messagebox.showinfo("완료", f"변경된 파일이 {destination_file}로 저장되었습니다.")
        print(f'gpio_test_row: {gpio_test_row}, peri_test_row: {peri_test_row}')  # 찾은 행 출력
    except PermissionError:
        messagebox.showerror("오류", f"파일을 저장할 수 없습니다. '{destination_file}' 파일이 열려 있는지 확인하세요.")

def process_xlsx_to_pattern_all_pin(file_path, pin_type=1, start_margin = 0, output_margin=0):   #pin_type : 0 = test pin / 1 = all pin
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    calculated_file_path = file_path.replace('.xlsx', '_calculated.xlsx')
    workbook.save(calculated_file_path)

    df = pd.read_excel(calculated_file_path, sheet_name='Sheet')
    sheet = workbook['Sheet']

    previous_u_color = None
    previous_v_color = None
    gpio_test_row = None  # 첫 번째 조건을 만족하는 행
    peri_test_row = None  # 두 번째 조건을 만족하는 행
    peri_test_count = 0   # peri_test_row를 찾기 위한 카운트, 초기화 필요
    gpio_output_row = 0

    try:
        cycle_time = int(entry_pattern_time_increment.get())
    except ValueError:
        messagebox.showerror("오류", "유효한 숫자를 입력하세요.")
        return

    for i, row in enumerate(df.itertuples(index=False), start=3):
        # D열이 1이고 E열에서 P열이 0인 첫 번째 행 찾기 (gpio_test_row)
        if gpio_test_row is None:
            if sheet.cell(row=i, column=4).value == 1 and all(sheet.cell(row=i, column=j).value == 0 for j in range(5, 17)):
                gpio_test_row = i

        # gpio_test_row 이후에 D열이 0이고 E열이 1이며 F열에서 P열이 0인 행을 카운트해서 두 번째 행을 찾기 (peri_test_row)
        if gpio_test_row is not None and i > gpio_test_row:
            if (sheet.cell(row=i, column=4).value == 0 and 
                sheet.cell(row=i, column=5).value == 1 and 
                all(sheet.cell(row=i, column=j).value == 0 for j in range(6, 17))):
                peri_test_count += 1  # 조건을 만족할 때마다 카운트 증가
                if peri_test_count == 2:  # 두 번째로 해당하는 행을 peri_test_row에 저장
                    peri_test_row = i

    if pin_type == 0:
        gpio_pattern_lines = []
        peri_pattern_lines = []

        if peri_test_row is not None:
            u_data_offset = sheet.cell(row=peri_test_row, column=21).value
            u_data_offset = round(u_data_offset) if not pd.isna(u_data_offset) else 0

        for index, row in df.iloc[2:].iterrows():
            u_data = row.iloc[20]
            v_data = row.iloc[21] 

            u_data = round(u_data) if not pd.isna(u_data) else 0
            v_data = round(v_data) if not pd.isna(v_data) else 0
            d_to_o_data = row.iloc[3:15].values.tolist()
            p_to_s_data = row.iloc[15:19].values.tolist()

            u_cell = sheet[f'U{index + 2}']
            u_color = u_cell.font.color.rgb if u_cell.font.color and u_cell.font.color.rgb else None
            v_cell = sheet[f'V{index + 2}']
            v_color = v_cell.font.color.rgb if v_cell.font.color and v_cell.font.color.rgb else None
  
            if gpio_test_row is not None and peri_test_row is not None and index + 2 >= peri_test_row:
                u_data -= u_data_offset
            
            u_data += int(start_margin)

            if u_color == 'FFFF0000' and v_color == 'FFFF0000':
                v_data = int(u_data/cycle_time)
                prefix = 'W "T1"; V { all_pin\t=\t'
                if gpio_test_row is not None and peri_test_row is not None and index + 2 < peri_test_row:
                    data = " ".join([str(int(x)) if x is not None else 'X' for x in d_to_o_data[:6]]) + " X X X X X X X X X X X X X X X X X "
                    suffix = f";}} // {v_data}, {u_data}ns"
                    line = prefix + data + suffix
                    gpio_pattern_lines.append(line)
                elif gpio_test_row is not None and peri_test_row is not None and index + 2 >= peri_test_row:
                    data = " ".join([str(int(x)) if x is not None else 'X' for x in d_to_o_data[:6]]) + " X X X X X X "
                    suffix = f";}} // {v_data}, {u_data}ns"
                    line = prefix + data + suffix
                    peri_pattern_lines.append(line)

            elif u_color == 'FF0000FF' and v_color == 'FF0000FF':
                # peri test output일 경우
                if index + 2 >= peri_test_row:
                    u_data_adjusted = u_data + int(output_margin)
                    v_data_adjusted = int(u_data_adjusted/cycle_time)
                    prefix = "W \"T1\"; V { all_pin\t=\t"
                    data = "X X X X X X " + " ".join(['H' if x == 1 else 'L' for x in d_to_o_data[-6:]]) + " "# + " X X X X X X X X X "
                    suffix = f";}} // {v_data_adjusted}, {u_data_adjusted}ns"
                    line = prefix + data + suffix
                    peri_pattern_lines.append(line)

                # gpio test output일 경우
                if gpio_test_row is not None and peri_test_row is not None and gpio_test_row < index + 2 < peri_test_row:
                    gpio_output_row += 1

                    if gpio_output_row < 3:
                        u_data_adjusted = u_data + int(output_margin)
                        v_data_adjusted = int(u_data_adjusted/cycle_time)
                        prefix = "W \"T1\"; V { all_pin\t=\t"
                        data = " ".join(['H' if x == 1 else 'L' for x in d_to_o_data[:6]]) + " L L L L L L L L L L L L L L L L L "
                        suffix = f";}} // {v_data_adjusted}, {u_data_adjusted}ns"
                        line = prefix + data + suffix
                        gpio_pattern_lines.append(line)

                    elif gpio_output_row == 3:
                        for k in range(23 - gpio_output_row + 1):
                            u_data_adjusted += 20000
                            v_data_adjusted = int(u_data_adjusted/cycle_time)
                            prefix = "W \"T1\"; V { all_pin\t=\t"
                            data_prefix = 'L ' * (k + 2) + 'H '
                            data_suffix = 'L ' * (23 - k - 3)
                            suffix = f";}} // {v_data_adjusted}, {u_data_adjusted}ns"
                            line = prefix + data_prefix + data_suffix + suffix
                            gpio_pattern_lines.append(line)

        gpio_pattern_file_path = file_path.replace('.xlsx', '_gpio_pattern.txt')
        peri_pattern_file_path = file_path.replace('.xlsx', '_peri_pattern.txt')

        with open(gpio_pattern_file_path, 'w') as file:
            for line in gpio_pattern_lines:
                file.write('\t' + line + '\n')
        with open(peri_pattern_file_path, 'w') as file:
            for line in peri_pattern_lines:
                file.write('\t' + line + '\n')
        messagebox.showinfo("완료", f"{gpio_pattern_file_path} {peri_pattern_file_path} 파일 작성이 완료되었습니다.")

    else:
        pattern_lines = []
        for index, row in df.iloc[2:].iterrows():
            u_data = row.iloc[20]
            v_data = row.iloc[21] 

            u_data = round(u_data) if not pd.isna(u_data) else 0
            v_data = round(v_data) if not pd.isna(v_data) else 0
            d_to_o_data = row.iloc[3:15].values.tolist()
            p_to_s_data = row.iloc[15:19].values.tolist()

            u_cell = sheet[f'U{index + 2}']
            u_color = u_cell.font.color.rgb if u_cell.font.color and u_cell.font.color.rgb else None
            v_cell = sheet[f'V{index + 2}']
            v_color = v_cell.font.color.rgb if v_cell.font.color and v_cell.font.color.rgb else None

            u_data += int(start_margin)
            
            if u_color == 'FFFF0000' and v_color == 'FFFF0000':
                v_data = int(u_data/cycle_time)
                prefix = 'W "T1"; V { all_pin\t=\t'
                data = " ".join([str(int(x)) if x is not None else 'X' for x in d_to_o_data[:6]]) + " X X X X X X X X X X X X X X X X X "
                suffix = f";}} // {v_data}, {u_data}ns"
                line = prefix + data + suffix
                pattern_lines.append(line)

            elif u_color == 'FF0000FF' and v_color == 'FF0000FF':
                #peri test output일 경우
                if index + 2 >= peri_test_row:
                    u_data_adjusted = u_data + int(output_margin)
                    v_data_adjusted = int(u_data_adjusted/cycle_time)
                    prefix = "W \"T1\"; V { all_pin\t=\t"
                    data = "X X X X X X X X " + " ".join(['H' if x == 1 else 'L' for x in d_to_o_data[-6:]]) + " X X X X X X X X X "
                    suffix = f";}} // {v_data_adjusted}, {u_data_adjusted}ns"
                    line = prefix + data + suffix
                    pattern_lines.append(line)
                # gpio test output일 경우
                if gpio_test_row is not None and peri_test_row is not None and gpio_test_row < index + 2 < peri_test_row:
                    gpio_output_row += 1

                    if gpio_output_row < 3:
                        u_data_adjusted = u_data + int(output_margin)
                        v_data_adjusted = int(u_data_adjusted/cycle_time)
                        prefix = "W \"T1\"; V { all_pin\t=\t"
                        data = " ".join(['H' if x == 1 else 'L' for x in d_to_o_data[:6]]) + " L L L L L L L L L L L L L L L L L "
                        suffix = f";}} // {v_data_adjusted}, {u_data_adjusted}ns"
                        line = prefix + data + suffix
                        pattern_lines.append(line)

                    elif gpio_output_row == 3:
                        for k in range(23 - gpio_output_row + 1):
                            u_data_adjusted += 20000
                            v_data_adjusted = int(u_data_adjusted/cycle_time)
                            prefix = "W \"T1\"; V { all_pin\t=\t"
                            data_prefix = 'L ' * (k + 2) + 'H '
                            data_suffix = 'L ' * (23 - k - 3)
                            suffix = f";}} // {v_data_adjusted}, {u_data_adjusted}ns"
                            line = prefix + data_prefix + data_suffix + suffix
                            pattern_lines.append(line)

        peri_pattern_file_path = file_path.replace('.xlsx', '_pattern.txt')
        with open(peri_pattern_file_path, 'w') as file:
            for line in pattern_lines:
                file.write('\t' + line + '\n')
        messagebox.showinfo("완료", f"{peri_pattern_file_path} 파일 작성이 완료되었습니다.")

    if os.path.exists(calculated_file_path):
        os.remove(calculated_file_path)
        print(f'{calculated_file_path} 파일이 삭제되었습니다.')
    else:
        print(f'{calculated_file_path} 파일을 찾을 수 없습니다.')

def fill_missing_patterns(file_path, time_increment):
    with open(file_path, 'r') as file:
        lines = file.readlines()

    new_lines = []
    step_count = 0

    stil_header_function = """
// Cycle time     : 5.000 nano seconds, ANTN[NRZ],ANTP[NRZ]
// Pattern depth  : 1750339
// Pattern order  : pin65,pin66,pin67,pin68,pin9,pin10,pin11,pin12,pin14,pin15,pin16,pin17,pin18,pin19,pin20,pin21,pin30,pin40,pin41,pin46,pin47,pin48,pin49

STIL 1.0;
Signals {
    GPIO0     In/Out;
    GPIO1     In/Out;
    GPIO2     In/Out;
    GPIO3     In/Out;
    GPIO4     In/Out;
    GPIO5     In/Out;
    GPIO6     In/Out;
    GPIO7     In/Out;
    GPIO8     In/Out;
    GPIO9     In/Out;
    GPIO10    In/Out;
    GPIO11    In/Out;
    GPIO12    In/Out;
    GPIO13    In/Out;
    GPIO14    In/Out;
    GPIO15    In/Out;
    GPIO16    In/Out;
    GPIO18    In/Out;
    GPIO19    In/Out;
    GPIO26    In/Out;
    GPIO27    In/Out;
    GPIO28    In/Out;
    GPIO29    In/Out;
}

SignalGroups {
    allpin = 'GPIO0+GPIO1+GPIO2+GPIO3+GPIO4+GPIO5+GPIO6+GPIO7+GPIO8+GPIO9+GPIO10+GPIO11+GPIO12+GPIO13+GPIO14+GPIO15+GPIO16+GPIO18+GPIO19+GPIO26+GPIO27+GPIO28+GPIO29';
}
//PPRO FUNCTION_TEST_R3B
//

//SUBSECT FUNCTION_TEST_R3B
Pattern FUNCTION_TEST_R3B {"""

    stil_header_peri = """
// Cycle time     : 5.000 nano seconds, ANTN[NRZ],ANTP[NRZ]
// Pattern depth  : 1750339
// Pattern order  : pin65,pin66,pin67,pin68,pin9,pin10,pin14,pin15,pin16,pin17,pin18,pin19

STIL 1.0;
Signals {
    GPIO0     In/Out;
    GPIO1     In/Out;
    GPIO2     In/Out;
    GPIO3     In/Out;
    GPIO4     In/Out;
    GPIO5     In/Out;
    GPIO8     In/Out;
    GPIO9     In/Out;
    GPIO10    In/Out;
    GPIO11    In/Out;
    GPIO12    In/Out;
    GPIO13    In/Out;
}

SignalGroups {
    allpin = 'GPIO0+GPIO1+GPIO2+GPIO3+GPIO4+GPIO5+GPIO8+GPIO9+GPIO10+GPIO11+GPIO12+GPIO13';
}
//PPRO PERI_TEST_R3B
//

//SUBSECT PERI_TEST_R3B
Pattern PERI_TEST_R3B {"""

    start_line_index = None  # 초기화
    start_line_time = None  # 초기화

    for i, line in enumerate(lines):
        if "//" in line:
            if start_line_index is None and start_line_time is None:
                start_line_index = int(line.split("//")[-1].strip().split(',')[0])
                start_line_time = int(line.split(",")[-1].strip().replace('ns', ''))

            first_tab_index = line.index("\t") + 1
            second_tab_index = line.index("\t", first_tab_index) + 1

            start_index = line.index("\t", second_tab_index) + 1
            first_colon_index = line.index(";") + 1
            colon_index = line.index(";", first_colon_index)

            # char_cnt로 GPIO 핀의 상태 개수를 확인
            char_cnt = len(line[start_index:colon_index].replace(" ", ""))

        if char_cnt >= 0:
            break

    print(f'start_line_index : {start_line_index} start_line_time : {start_line_time}')

    if char_cnt > 12:
        # FUNCTION_TEST_R3B 형식
        if not new_lines or "Pattern FUNCTION_TEST_R3B" not in new_lines[0]:
            new_lines.append(stil_header_function)  # 처음 한 번 헤더 추가
    else:
        # PERI_TEST_R3B 형식
        if not new_lines or "Pattern PERI_TEST_R3B" not in new_lines[0]:
            new_lines.append(stil_header_peri)  # 처음 한 번 헤더 추가          

    current_index = 0
    current_time = 0

    if start_line_index > 0:
        while current_index + 1 < start_line_index:
            step_count += 1
            x_values = "X " * char_cnt
            new_line = f'W "T1"; V {{ all_pin\t=\t{x_values.strip()} ;}} // {current_index}, {current_time}ns'
            new_lines.append(new_line)
            current_index += 1
            current_time += time_increment  # 시간 증가


    for i, line in enumerate(lines):
        new_lines.append(line.strip())
        
        if "//" in line:
            current_index = int(line.split("//")[-1].strip().split(',')[0])
            current_time = int(line.split(",")[-1].strip().replace('ns', ''))
            
            if i + 1 < len(lines) and "//" in lines[i + 1]:
                next_index = int(lines[i + 1].split("//")[-1].strip().split(',')[0])
                next_time = int(lines[i + 1].split(",")[-1].strip().replace('ns', ''))

                while current_index + 1 < next_index:
                    step_count += 1
                    x_values = "X " * char_cnt
                    current_index += 1
                    current_time += time_increment  # 시간 증가
                    new_line = f'W "T1"; V {{ all_pin\t=\t{x_values.strip()} ;}} // {current_index}, {current_time}ns'
                    new_lines.append(new_line)

    # STIL 형식의 패턴 종료 부분 추가
    stil_footer = '}'
    new_lines.append(stil_footer)

    # 새로운 파일을 저장
    base_name, ext = os.path.splitext(file_path)
    new_file_path = f"{base_name}_total{ext}"

    with open(new_file_path, 'w') as file:
        for line in new_lines:
            if line == '}':
                file.write(line + '\n')  # 탭 없이 종료 구문 추가
            else:
                file.write('\t' + line + '\n')  # 나머지는 탭 추가

    messagebox.showinfo("완료", f"{new_file_path} 파일 작성이 완료되었습니다.")

def run_xlsx_process_all_pin():
    try:
        time_increment = int(entry_csv_time_increment.get())
    except ValueError:
        messagebox.showerror("오류", "유효한 숫자를 입력하세요.")
        return
    file_path = entry_csv_file_path.get()
    if file_path:
        process_csv_to_xlsx_all_pin(file_path, time_increment)
    else:
        messagebox.showerror("오류", "파일 경로를 선택하세요.")

def run_pattern_process_separate():
    file_path = entry_pattern_file_path.get()
    pin_type = 0

    try:
        start_margin = float(entry_pattern_start_margin.get())
    except ValueError:
        start_margin = 0

    try:
        output_margin = float(entry_pattern_output_margin.get())
    except ValueError:
        output_margin = 0

    if file_path:
        process_xlsx_to_pattern_all_pin(file_path, pin_type, start_margin, output_margin)
    else:
        messagebox.showerror("오류", "파일 경로를 선택하세요.")

def run_pattern_process_single():
    file_path = entry_pattern_file_path.get()
    pin_type = 1

    try:
        start_margin = float(entry_pattern_start_margin.get())
    except ValueError:
        start_margin = 0

    try:
        output_margin = float(entry_pattern_output_margin.get())
    except ValueError:
        output_margin = 0

    if file_path:
        process_xlsx_to_pattern_all_pin(file_path, pin_type, start_margin, output_margin)
    else:
        messagebox.showerror("오류", "파일 경로를 선택하세요.")

def run_xlsx_process_based_on_selection():
    run_xlsx_process_all_pin()

def run_pattern_process_based_on_selection():
    if pattern_var.get() == "separate":
        run_pattern_process_separate()
    else:
        run_pattern_process_single()       
        
def run_total_pattern_process():
    file_path = entry_txt_file_path.get()
    try:
        time_increment = int(entry_time_increment.get())
    except ValueError:
        messagebox.showerror("오류", "유효한 숫자를 입력하세요.")
        return
    if file_path:
        fill_missing_patterns(file_path, time_increment)
    else:
        messagebox.showerror("오류", "파일 경로를 선택하세요.")

def select_file(entry, file_type="csv"):
    file_path = filedialog.askopenfilename(filetypes=[(f"{file_type.upper()} files", f"*.{file_type}")])
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)

# GUI 생성
root = tk.Tk()
root.title("Pattern Generator")

label_csv_name = tk.Label(root, text="[csv to xlsx]")
label_csv_name.grid(row=1, column=0, padx=2, pady=10)
label_csv_file_path = tk.Label(root, text="File path:")
label_csv_file_path.grid(row=1, column=3, padx=5, pady=10)
entry_csv_file_path = tk.Entry(root, width=30)
entry_csv_file_path.grid(row=1, column=4, padx=5, pady=10)
button_csv_browse = tk.Button(root, text="search", command=lambda: select_file(entry_csv_file_path, "csv"))
button_csv_browse.grid(row=1, column=5, padx=5, pady=10)
label_csv_time_increment = tk.Label(root, text="Cycle(ns):")
label_csv_time_increment.grid(row=1, column=8, padx=5, pady=10)
entry_csv_time_increment = tk.Entry(root, width=10)
entry_csv_time_increment.grid(row=1, column=9, padx=5, pady=10)
entry_csv_time_increment.insert(0, "5")
button_csv_run = tk.Button(root, text="실행", command=run_xlsx_process_based_on_selection)
button_csv_run.grid(row=1, column=12, padx=10, pady=10)

# pattern 만들기 부분
pattern_var = tk.StringVar(value="single")  # 기본값을 "all"로 설정
label_pattern_name = tk.Label(root, text="[xlsx to pattern]")
label_pattern_name.grid(row=2, column=0, padx=2, pady=10)
rb_all_pattern = tk.Radiobutton(root, text="Single", variable=pattern_var, value="single")
rb_all_pattern.grid(row=2, column=1, padx=2, pady=10)
rb_test_pattern = tk.Radiobutton(root, text="Separate", variable=pattern_var, value="separate")
rb_test_pattern.grid(row=2, column=2, padx=2, pady=10)
label_pattern_file_path = tk.Label(root, text="File path:")
label_pattern_file_path.grid(row=2, column=3, padx=5, pady=10)
entry_pattern_file_path = tk.Entry(root, width=30)
entry_pattern_file_path.grid(row=2, column=4, padx=5, pady=10)
button_pattern_browse = tk.Button(root, text="search", command=lambda: select_file(entry_pattern_file_path, "xlsx"))
button_pattern_browse.grid(row=2, column=5, padx=5, pady=10)
label_pattern_start_margin = tk.Label(root, text="Start Margin:")
label_pattern_start_margin.grid(row=2, column=6, padx=5, pady=10)
entry_pattern_start_margin = tk.Entry(root, width=10)
entry_pattern_start_margin.grid(row=2, column=7, padx=5, pady=10)
entry_pattern_start_margin.insert(0, "0")
label_pattern_output_margin = tk.Label(root, text="Output Margin:")
label_pattern_output_margin.grid(row=2, column=8, padx=5, pady=10)
entry_pattern_output_margin = tk.Entry(root, width=10)
entry_pattern_output_margin.grid(row=2, column=9, padx=5, pady=10)
entry_pattern_output_margin.insert(0, "0")
label_pattern_time_increment = tk.Label(root, text="Cycle(ns):")
label_pattern_time_increment.grid(row=2, column=10, padx=5, pady=10)
entry_pattern_time_increment = tk.Entry(root, width=10)
entry_pattern_time_increment.grid(row=2, column=11, padx=5, pady=10)
entry_pattern_time_increment.insert(0, "5")
button_pattern_run = tk.Button(root, text="실행", command=run_pattern_process_based_on_selection)
button_pattern_run.grid(row=2, column=12, padx=10, pady=10)

## 패턴 채우기
label_name = tk.Label(root, text="[Fill mask Pattern]")
label_name.grid(row=3, column=0, padx=2, pady=10)
label_file_path = tk.Label(root, text="File path:")
label_file_path.grid(row=3, column=3, padx=5, pady=10)
entry_txt_file_path = tk.Entry(root, width=30)
entry_txt_file_path.grid(row=3, column=4, padx=5, pady=10)
button_browse = tk.Button(root, text="search", command=lambda: select_file(entry_txt_file_path, "txt"))
button_browse.grid(row=3, column=5, padx=5, pady=10)
label_time_increment = tk.Label(root, text="Cycle(ns):")
label_time_increment.grid(row=3, column=8, padx=5, pady=10)
entry_time_increment = tk.Entry(root, width=10)
entry_time_increment.grid(row=3, column=9, padx=5, pady=10)
entry_time_increment.insert(0, "5")
button_run = tk.Button(root, text="실행", command=run_total_pattern_process)
button_run.grid(row=3, column=12, padx=10, pady=10)

root.mainloop()